"""Excel & PDF export utilities with Vietnamese-safe fonts."""
from io import BytesIO
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont


# Register Vietnamese-safe fonts (DejaVu Sans has full Latin Extended-A coverage)
_FONT_DIR = Path(__file__).parent / "fonts"
_FONT_REGISTERED = False


def _register_fonts():
    global _FONT_REGISTERED
    if _FONT_REGISTERED:
        return
    reg = _FONT_DIR / "DejaVuSans.ttf"
    bold = _FONT_DIR / "DejaVuSans-Bold.ttf"
    if reg.exists():
        pdfmetrics.registerFont(TTFont("MKG", str(reg)))
        if bold.exists():
            pdfmetrics.registerFont(TTFont("MKG-Bold", str(bold)))
        else:
            pdfmetrics.registerFont(TTFont("MKG-Bold", str(reg)))
    else:
        # Fall back to Helvetica (limited Unicode)
        pdfmetrics.registerFont(TTFont("MKG", str(reg)) if reg.exists() else None)
    _FONT_REGISTERED = True


def build_excel(title: str, headers: list, rows: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "BaoCao"
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=14, color="00A82D")
    ws.append([])
    ws.append(headers)
    for cell in ws[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="00A3E0")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for r in rows:
        ws.append(r)
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(40, max(12, max_len + 2))
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def build_pdf(title: str, headers: list, rows: list) -> bytes:
    _register_fonts()
    bio = BytesIO()
    doc = SimpleDocTemplate(
        bio, pagesize=landscape(A4),
        leftMargin=1.2 * cm, rightMargin=1.2 * cm,
        topMargin=1.2 * cm, bottomMargin=1.2 * cm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "MKGTitle", parent=styles["Title"],
        fontName="MKG-Bold", fontSize=16,
        textColor=colors.HexColor("#00A82D"),
        alignment=0,
    )
    subtitle_style = ParagraphStyle(
        "MKGSubtitle", parent=styles["Heading3"],
        fontName="MKG-Bold", fontSize=11,
        textColor=colors.HexColor("#0B1120"),
        alignment=0,
    )
    story = [
        Paragraph("MekongGreen — Hệ Thống Bản Đồ Số Cơ Giới Hóa Nông Nghiệp", title_style),
        Paragraph(title, subtitle_style),
        Spacer(1, 0.4 * cm),
    ]
    data = [headers] + [[str(c) if c is not None else "" for c in r] for r in rows]
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#00A3E0")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "MKG-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "MKG"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E2E8F0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(table)
    doc.build(story)
    return bio.getvalue()
