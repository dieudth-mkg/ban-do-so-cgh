"""MekongGreen - FastAPI backend."""
from fastapi import FastAPI, APIRouter, HTTPException, Depends, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pymongo import ReturnDocument
import os
import logging
from pathlib import Path
from typing import Optional, List
from io import BytesIO
from datetime import datetime, timezone
from openpyxl import load_workbook, Workbook
import httpx
import random
import uuid

from models import (
    UserCreate, LoginRequest, ChangePasswordRequest,
    HTX, HTXCreate, Machine, MachineCreate, MachineHistory,
    Owner, OwnerCreate,
    MachineCategory, MachineCategoryCreate,
    ProductivityNorm, ProductivityNormCreate,
    Province, ProvinceUpdate,
    ProductionStage, ProductionStageCreate,
    SeasonType, SeasonTypeCreate,
    SystemParams,
)
from auth import (
    hash_password, verify_password, create_token,
    get_current_user, require_admin,
)
from seed import seed_all, PROVINCES, STAGES, SEASONS, MACHINE_CATEGORIES, NORMS
from exports import build_excel, build_pdf, build_excel_by_region, build_pdf_by_region
from geojson_data import PROVINCE_GEOJSON


# Season demand factor - applied to cultivated_area_ha when computing needed machines
SEASON_FACTOR = {"DX": 1.0, "HT": 0.9, "TD": 0.6, "ALL": 1.0}

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

mongo_url = os.environ["MONGO_URL"]
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ["DB_NAME"]]

app = FastAPI(title="MekongGreen API")
api = APIRouter(prefix="/api")

logger = logging.getLogger("mekonggreen")
logging.basicConfig(level=logging.INFO)


# ============ FN-04 HELPERS ============
def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


async def _log_action(user: dict, action: str, detail: str = ""):
    """FN-12/BR-01: mọi thao tác Thêm/Sửa/Vô hiệu hóa dữ liệu đều được ghi nhật ký."""
    await db.system_logs.insert_one({
        "id": f"log-{datetime.now(timezone.utc).timestamp()}",
        "actor_email": (user or {}).get("email", "system"),
        "action": action,
        "detail": detail,
        "ts": _now_iso(),
    })


SOURCE_LABEL = {"app_htx": "[App HTX]", "manual": "[Nhập tay]", "no_data": "[Chưa có dữ liệu]"}
# Chỉ máy đang active (chưa vô hiệu hóa) mới được tính năng lực (BR-05/BR-11)
ACTIVE_FILTER = {"active": {"$ne": False}}


async def _next_machine_code() -> str:
    """BR-01: Mã máy do hệ thống TỰ SINH, duy nhất toàn hệ thống."""
    doc = await db.counters.find_one_and_update(
        {"_id": "machine_code"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=ReturnDocument.AFTER,
    )
    return f"MAY-{doc['seq']:05d}"


async def _next_owner_code() -> str:
    """FN-03: Mã CSH do hệ thống TỰ SINH."""
    doc = await db.counters.find_one_and_update(
        {"_id": "owner_code"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=ReturnDocument.AFTER,
    )
    return f"CSH-{doc['seq']:05d}"


async def _find_machine_by_identity(serial_no: str, chassis_no: str, exclude_id: str = None):
    """BR-01: đối chiếu theo Số máy/SN, thiếu SN thì dùng Số khung."""
    ors = []
    if (serial_no or "").strip():
        ors.append({"serial_no": serial_no.strip()})
    if (chassis_no or "").strip():
        ors.append({"chassis_no": chassis_no.strip()})
    if not ors:
        return None
    q = {"$or": ors}
    if exclude_id:
        q["id"] = {"$ne": exclude_id}
    return await db.machines.find_one(q)


async def _add_machine_history(machine_id, machine_code, change_type, source, actor,
                               field="", before=None, after=None,
                               owned_since=None, deactivated_at=None):
    """BR-07: ghi MỘT bản ghi mới vào Lịch sử biến động, không ghi đè bản cũ."""
    entry = MachineHistory(
        machine_id=machine_id, machine_code=machine_code or "",
        change_type=change_type, source=source, actor=actor or "",
        field=field,
        before=None if before is None else str(before),
        after=None if after is None else str(after),
        owned_since=owned_since, deactivated_at=deactivated_at,
    ).model_dump()
    await db.machine_history.insert_one(entry)


def _coerce_date(v) -> str:
    """Chuẩn hóa Ngày HTX sở hữu về chuỗi ISO (chấp nhận datetime/str)."""
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    return str(v).strip()


# ============ HEALTH ============
@api.get("/")
async def root():
    return {"service": "MekongGreen", "status": "ok"}


# ============ AUTH ============
@api.post("/auth/login")
async def login(payload: LoginRequest):
    doc = await db.users.find_one({"email": payload.email})
    if not doc or not doc.get("active", True):
        raise HTTPException(status_code=401, detail="Email hoặc mật khẩu không đúng")
    if not verify_password(payload.password, doc["password_hash"]):
        raise HTTPException(status_code=401, detail="Email hoặc mật khẩu không đúng")
    token = create_token(doc["id"], doc["email"], doc["role"])
    await db.system_logs.insert_one({
        "id": f"log-{datetime.now(timezone.utc).timestamp()}",
        "actor_email": doc["email"],
        "action": "LOGIN",
        "detail": "",
        "ts": datetime.now(timezone.utc).isoformat(),
    })
    return {
        "token": token,
        "user": {
            "id": doc["id"],
            "email": doc["email"],
            "full_name": doc["full_name"],
            "role": doc["role"],
            "must_change_password": doc.get("must_change_password", False),
        },
    }


@api.get("/auth/me")
async def me(user=Depends(get_current_user)):
    doc = await db.users.find_one({"email": user["email"]}, {"_id": 0, "password_hash": 0})
    return doc


@api.post("/auth/change-password")
async def change_password(payload: ChangePasswordRequest, user=Depends(get_current_user)):
    doc = await db.users.find_one({"email": user["email"]})
    if not doc or not verify_password(payload.old_password, doc["password_hash"]):
        raise HTTPException(status_code=400, detail="Mật khẩu cũ không đúng")
    await db.users.update_one(
        {"email": user["email"]},
        {"$set": {"password_hash": hash_password(payload.new_password), "must_change_password": False}},
    )
    return {"ok": True}


# ============ FN-01 NHÓM 3: ĐƠN VỊ HÀNH CHÍNH (Tỉnh) ============
@api.get("/provinces")
async def list_provinces():
    docs = await db.provinces.find({}, {"_id": 0}).to_list(200)
    return docs


@api.post("/provinces")
async def create_province(body: Province, user=Depends(require_admin)):
    existing = await db.provinces.find_one({"code": body.code})
    if existing:
        raise HTTPException(status_code=400, detail="Mã đơn vị đã tồn tại")
    doc = body.model_dump()
    await db.provinces.insert_one(doc)
    await _log_action(user, "CREATE_PROVINCE", f"code={body.code}, name={body.name}")
    return {k: v for k, v in doc.items() if k != "_id"}


@api.patch("/provinces/{code}")
async def update_province(code: str, body: ProvinceUpdate, user=Depends(require_admin)):
    """BR-02: đổi địa giới → đóng hiệu lực bản ghi cũ (active=False), KHÔNG xóa cứng."""
    allowed = {k: v for k, v in body.model_dump(exclude_none=True).items()}
    if not allowed:
        return {"ok": True}
    await db.provinces.update_one({"code": code}, {"$set": allowed})
    await _log_action(user, "UPDATE_PROVINCE", f"code={code}, fields={list(allowed.keys())}")
    return {"ok": True}


# ============ FN-01 NHÓM 1: KHÂU SẢN XUẤT ============
@api.get("/stages")
async def list_stages():
    return await db.stages.find({}, {"_id": 0}).to_list(50)


@api.post("/stages")
async def create_stage(body: ProductionStageCreate, user=Depends(require_admin)):
    existing = await db.stages.find_one({"code": body.code})
    if existing:
        raise HTTPException(status_code=400, detail="Mã khâu sản xuất đã tồn tại")
    doc = ProductionStage(**body.model_dump()).model_dump()
    await db.stages.insert_one(doc)
    await _log_action(user, "CREATE_STAGE", f"code={body.code}, name={body.name}")
    return {k: v for k, v in doc.items() if k != "_id"}


@api.patch("/stages/{code}")
async def update_stage(code: str, body: dict, user=Depends(require_admin)):
    allowed = {k: v for k, v in body.items() if k in ("name", "active", "pilot_enabled")}
    await db.stages.update_one({"code": code}, {"$set": allowed})
    await _log_action(user, "UPDATE_STAGE", f"code={code}, fields={list(allowed.keys())}")
    return {"ok": True}


# ============ FN-01 NHÓM 2: TÊN VỤ SẢN XUẤT ============
@api.get("/seasons")
async def list_seasons():
    return await db.seasons.find({}, {"_id": 0}).to_list(50)


@api.post("/seasons")
async def create_season(body: SeasonTypeCreate, user=Depends(require_admin)):
    existing = await db.seasons.find_one({"code": body.code})
    if existing:
        raise HTTPException(status_code=400, detail="Mã vụ đã tồn tại")
    doc = SeasonType(**body.model_dump()).model_dump()
    await db.seasons.insert_one(doc)
    await _log_action(user, "CREATE_SEASON", f"code={body.code}, name={body.name}")
    return {k: v for k, v in doc.items() if k != "_id"}


@api.patch("/seasons/{code}")
async def update_season(code: str, body: dict, user=Depends(require_admin)):
    allowed = {k: v for k, v in body.items() if k in ("name", "active")}
    await db.seasons.update_one({"code": code}, {"$set": allowed})
    await _log_action(user, "UPDATE_SEASON", f"code={code}, fields={list(allowed.keys())}")
    return {"ok": True}


# ============ MACHINE CATEGORIES ============
@api.get("/machine-categories")
async def list_categories(user=Depends(get_current_user)):
    return await db.machine_categories.find({}, {"_id": 0}).to_list(200)


@api.post("/machine-categories")
async def create_category(body: MachineCategoryCreate, user=Depends(require_admin)):
    existing = await db.machine_categories.find_one({"code": body.code})
    if existing:
        raise HTTPException(status_code=400, detail="Mã chủng loại đã tồn tại")
    doc = MachineCategory(**body.model_dump()).model_dump()
    await db.machine_categories.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}


@api.patch("/machine-categories/{code}")
async def update_category(code: str, body: dict, user=Depends(require_admin)):
    allowed = {k: v for k, v in body.items() if k in ("name", "stage", "active")}
    await db.machine_categories.update_one({"code": code}, {"$set": allowed})
    return {"ok": True}


# ============ FN-01 NHÓM 4: ĐỊNH MỨC DIỆN TÍCH/MÁY/VỤ ============
@api.get("/productivity-norms")
async def list_norms(user=Depends(get_current_user)):
    return await db.productivity_norms.find({}, {"_id": 0}).to_list(200)


@api.post("/productivity-norms")
async def upsert_norm(body: ProductivityNormCreate, user=Depends(require_admin)):
    """BR-01/BR-04: định mức là DỮ LIỆU do DCRD ban hành bằng văn bản, khoá theo Chủng loại × Khâu,
    bắt buộc có ngày hiệu lực và số/ngày văn bản DCRD."""
    payload = body.model_dump()
    payload["effective_from"] = payload.get("effective_from") or datetime.now(timezone.utc).isoformat()
    payload["active"] = True
    key = {"category_code": body.category_code, "stage_code": body.stage_code}
    await db.productivity_norms.update_one(key, {"$set": payload}, upsert=True)
    await _log_action(
        user, "UPDATE_NORM",
        f"category={body.category_code}, stage={body.stage_code}, "
        f"ha/machine={body.ha_per_machine_per_season}, doc={body.document_ref}",
    )
    return {"ok": True}


@api.patch("/productivity-norms/{norm_id}")
async def update_norm_status(norm_id: str, body: dict, user=Depends(require_admin)):
    """BR-03: không xóa cứng, chỉ ngừng hiệu lực (active=False) hoặc đóng effective_to."""
    allowed = {k: v for k, v in body.items() if k in ("active", "effective_to", "document_ref")}
    await db.productivity_norms.update_one({"id": norm_id}, {"$set": allowed})
    await _log_action(user, "UPDATE_NORM_STATUS", f"id={norm_id}, fields={list(allowed.keys())}")
    return {"ok": True}


# ============ FN-01 NHÓM 5: NGƯỠNG CẢNH BÁO CUNG-CẦU (đầy đủ 4 mốc) ============
@api.get("/admin/thresholds")
async def get_thresholds_v2(user=Depends(get_current_user)):
    doc = await db.alert_thresholds.find_one({"key": "default"}, {"_id": 0})
    return doc or {"sufficient_min": 0.85, "slight_min": 0.60, "excess_min": 1.20}


@api.patch("/admin/thresholds")
async def update_thresholds_v2(body: dict, user=Depends(require_admin)):
    allowed = {}
    for k in ("sufficient_min", "slight_min", "excess_min"):
        if k in body:
            allowed[k] = float(body[k])
    allowed["effective_from"] = datetime.now(timezone.utc).isoformat()
    await db.alert_thresholds.update_one({"key": "default"}, {"$set": allowed}, upsert=True)
    await _log_action(user, "UPDATE_THRESHOLDS", f"fields={list(allowed.keys())}")
    return {"ok": True}


# ============ FN-01 NHÓM 6: THAM SỐ HỆ THỐNG ============
@api.get("/admin/system-params")
async def get_system_params(user=Depends(get_current_user)):
    doc = await db.system_params.find_one({"key": "default"}, {"_id": 0})
    return doc or SystemParams().model_dump()


@api.patch("/admin/system-params")
async def update_system_params(body: dict, user=Depends(require_admin)):
    allowed = {k: v for k, v in body.items() if k in (
        "basemap_source", "import_allowed_extensions", "import_max_rows",
        "import_max_file_size_mb", "data_freshness_threshold_hours",
    )}
    await db.system_params.update_one({"key": "default"}, {"$set": allowed}, upsert=True)
    await _log_action(user, "UPDATE_SYSTEM_PARAMS", f"fields={list(allowed.keys())}")
    return {"ok": True}


# ============ HTX ============
@api.get("/htx")
async def list_htx(
    province: Optional[str] = None,
    q: Optional[str] = None,
    user=Depends(get_current_user),
):
    query = {}
    if province and province != "ALL":
        query["province_code"] = province
    if q:
        query["$or"] = [
            {"name": {"$regex": q, "$options": "i"}},
            {"code": {"$regex": q, "$options": "i"}},
            {"owner_name": {"$regex": q, "$options": "i"}},
        ]
    docs = await db.htx.find(query, {"_id": 0}).to_list(2000)
    return docs


@api.post("/htx")
async def create_htx(body: HTXCreate, user=Depends(require_admin)):
    if await db.htx.find_one({"code": body.code}):
        raise HTTPException(status_code=400, detail="Mã HTX đã tồn tại")
    doc = HTX(**body.model_dump()).model_dump()
    doc["id"] = doc["code"]
    await db.htx.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}


@api.patch("/htx/{code}")
async def update_htx(code: str, body: dict, user=Depends(require_admin)):
    allowed = {k: v for k, v in body.items() if k in (
        "name", "owner_name", "owner_type", "province_code", "district",
        "commune", "lat", "lng", "cultivated_area_ha", "phone", "active"
    )}
    result = await db.htx.update_one({"code": code}, {"$set": allowed})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Không tìm thấy HTX")
    return {"ok": True}


@api.delete("/htx/{code}")
async def deactivate_htx(code: str, user=Depends(require_admin)):
    await db.htx.update_one({"code": code}, {"$set": {"active": False}})
    return {"ok": True}


# ============ HTX EXCEL IMPORT ============
HTX_IMPORT_COLUMNS = [
    "code", "name", "owner_name", "province_code",
    "district", "commune", "lat", "lng", "cultivated_area_ha", "phone",
]


@api.get("/htx/import-template")
async def htx_import_template(user=Depends(require_admin)):
    """Download an Excel template for HTX bulk import."""
    wb = Workbook()
    ws = wb.active
    ws.title = "HTX_Template"
    ws.append(HTX_IMPORT_COLUMNS)
    # example rows
    ws.append(["CT-HTX99", "HTX Mẫu", "Nguyễn Văn A", "CT", "Ô Môn", "Thới An", 10.10, 105.60, 850, "0901234567"])
    ws.append(["AG-HTX99", "HTX Mẫu An Giang", "Trần Thị B", "AG", "Châu Đốc", "Vĩnh Mỹ", 10.70, 105.10, 1200, "0912345678"])
    for i, col in enumerate(ws.columns, 1):
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max(14, max_len + 2)
    bio = BytesIO(); wb.save(bio)
    return StreamingResponse(
        BytesIO(bio.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="htx-import-template.xlsx"'},
    )


@api.post("/htx/import-excel")
async def import_htx_excel(
    file: UploadFile = File(...),
    dry_run: bool = Query(False),
    user=Depends(require_admin),
):
    """Bulk import HTX from an .xlsx file. Returns row-level validation report."""
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Chỉ chấp nhận tệp .xlsx")
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Tệp vượt quá 10MB")
    try:
        wb = load_workbook(BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Không đọc được tệp Excel: {e}")
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Tệp trống hoặc chỉ có tiêu đề")
    header = [str(c).strip() if c is not None else "" for c in rows[0]]

    # map column index
    col_idx = {name: (header.index(name) if name in header else -1) for name in HTX_IMPORT_COLUMNS}
    missing = [k for k, v in col_idx.items() if v == -1 and k in ("code", "name", "owner_name", "province_code", "lat", "lng", "cultivated_area_ha")]
    if missing:
        raise HTTPException(status_code=400, detail=f"Thiếu cột bắt buộc: {', '.join(missing)}")

    prov_codes = {p["code"] for p in PROVINCES}
    ok_rows, error_rows, skipped_rows = [], [], []

    for r_i, r in enumerate(rows[1:], start=2):
        if all(c is None or str(c).strip() == "" for c in r):
            continue
        record = {}
        errors = []
        for k, idx in col_idx.items():
            record[k] = r[idx] if idx >= 0 and idx < len(r) else None

        # required
        for f in ("code", "name", "owner_name", "province_code", "lat", "lng", "cultivated_area_ha"):
            if record.get(f) in (None, ""):
                errors.append(f"Thiếu {f}")
        # province valid
        if record.get("province_code") and str(record["province_code"]).strip() not in prov_codes:
            errors.append(f"province_code không hợp lệ (chấp nhận: {', '.join(sorted(prov_codes))})")
        # numeric
        try:
            lat = float(record["lat"]); lng = float(record["lng"])
            if not (8.0 <= lat <= 24.0) or not (102.0 <= lng <= 110.0):
                errors.append("lat/lng ngoài phạm vi Việt Nam")
            record["lat"], record["lng"] = lat, lng
        except (TypeError, ValueError):
            errors.append("lat/lng phải là số")
        try:
            area = float(record["cultivated_area_ha"])
            if area <= 0:
                errors.append("cultivated_area_ha phải > 0")
            record["cultivated_area_ha"] = area
        except (TypeError, ValueError):
            errors.append("cultivated_area_ha phải là số")

        code = str(record.get("code") or "").strip()
        if not errors and code:
            existing = await db.htx.find_one({"code": code})
            if existing:
                skipped_rows.append({"row": r_i, "code": code, "reason": "Đã tồn tại"})
                continue

        if errors:
            error_rows.append({"row": r_i, "code": code, "errors": errors})
        else:
            record["code"] = code
            record["name"] = str(record["name"]).strip()
            record["owner_name"] = str(record["owner_name"]).strip()
            record["province_code"] = str(record["province_code"]).strip()
            record["district"] = str(record.get("district") or "").strip()
            record["commune"] = str(record.get("commune") or "").strip()
            record["phone"] = str(record.get("phone") or "").strip()
            ok_rows.append({"row": r_i, **record})

    inserted = 0
    if not dry_run and ok_rows:
        for rec in ok_rows:
            doc = {
                "id": rec["code"],
                "code": rec["code"],
                "name": rec["name"],
                "owner_name": rec["owner_name"],
                "owner_type": "HTX",
                "province_code": rec["province_code"],
                "district": rec["district"],
                "commune": rec["commune"],
                "lat": rec["lat"],
                "lng": rec["lng"],
                "cultivated_area_ha": rec["cultivated_area_ha"],
                "phone": rec["phone"],
                "active": True,
                "created_at": datetime.now(timezone.utc).isoformat(),
            }
            await db.htx.insert_one(doc)
            inserted += 1
        await db.system_logs.insert_one({
            "id": f"log-{datetime.now(timezone.utc).timestamp()}",
            "actor_email": user["email"],
            "action": "IMPORT_HTX_EXCEL",
            "detail": f"file={file.filename}, inserted={inserted}, errors={len(error_rows)}, skipped={len(skipped_rows)}",
            "ts": datetime.now(timezone.utc).isoformat(),
        })

    return {
        "filename": file.filename,
        "total_rows": len(rows) - 1,
        "ok_count": len(ok_rows),
        "error_count": len(error_rows),
        "skipped_count": len(skipped_rows),
        "inserted": inserted,
        "dry_run": dry_run,
        "errors": error_rows[:200],
        "skipped": skipped_rows[:200],
        "ok_preview": ok_rows[:20],
    }


# ============ GEOJSON PROVINCES ============
@api.get("/geojson/provinces")
async def geojson_provinces():
    """Return simplified GeoJSON polygons for the 6 provinces."""
    return PROVINCE_GEOJSON


# ============ HEATMAP HP/ha ============
@api.get("/map/heatmap")
async def map_heatmap(user=Depends(get_current_user)):
    """Return heat points [lat, lng, intensity] where intensity = HP density (HP/ha)."""
    htx_docs = await db.htx.find({"active": True}, {"_id": 0}).to_list(2000)
    machines = await db.machines.find({}, {"_id": 0}).to_list(20000)
    hp_by_htx: dict = {}
    for m in machines:
        hp_by_htx[m["htx_id"]] = hp_by_htx.get(m["htx_id"], 0) + m.get("horsepower", 0)
    points = []
    max_density = 0.0
    for h in htx_docs:
        area = h.get("cultivated_area_ha", 0)
        if not area:
            continue
        density = hp_by_htx.get(h["id"], 0) / area  # HP per ha
        max_density = max(max_density, density)
        points.append({"lat": h["lat"], "lng": h["lng"], "hp_per_ha": round(density, 3), "htx_code": h["code"]})
    return {"points": points, "max_density": round(max_density, 3)}


# ============ MACHINES ============
@api.get("/machines")
async def list_machines(
    htx_id: Optional[str] = None,
    owner_id: Optional[str] = None,
    category: Optional[str] = None,
    status: Optional[str] = None,
    province: Optional[str] = None,
    include_inactive: bool = False,
    user=Depends(get_current_user),
):
    query = {}
    if not include_inactive:               # BR-05: mặc định ẩn máy đã vô hiệu hóa
        query.update(ACTIVE_FILTER)
    if htx_id:
        query["htx_id"] = htx_id
    if owner_id:
        query["owner_id"] = owner_id
    if category and category != "ALL":
        query["category_code"] = category
    if status and status != "ALL":
        query["status"] = status
    if province and province != "ALL":
        htx_ids = [h["id"] async for h in db.htx.find({"province_code": province}, {"id": 1})]
        query["htx_id"] = {"$in": htx_ids}
    docs = await db.machines.find(query, {"_id": 0}).to_list(5000)
    # BR-02: khâu suy ra từ chủng loại; BR-04: nhãn nguồn tình trạng
    cat_stage = {c["code"]: c["stage"] for c in await db.machine_categories.find({}, {"_id": 0}).to_list(50)}
    for d in docs:
        d["stage"] = cat_stage.get(d.get("category_code"))
        d["status_source_label"] = SOURCE_LABEL.get(d.get("status_source", "manual"), "[Nhập tay]")
    return docs


@api.post("/machines")
async def create_machine(body: MachineCreate, user=Depends(require_admin)):
    htx = await db.htx.find_one({"id": body.htx_id})
    if not htx:
        raise HTTPException(status_code=404, detail="HTX không tồn tại")
    # BR-02: chủng loại phải khớp danh mục chuẩn
    cat = await db.machine_categories.find_one({"code": body.category_code})
    if not cat:
        raise HTTPException(status_code=400, detail="Chủng loại máy không hợp lệ")
    # BR-08: Ngày HTX sở hữu bắt buộc
    owned_since = _coerce_date(body.owned_since)
    if not owned_since:
        raise HTTPException(status_code=400, detail="Thiếu Ngày HTX sở hữu/tiếp nhận máy")
    # BR-01: định danh vật lý (Số máy/SN hoặc Số khung) — bắt buộc & duy nhất
    if not (body.serial_no or "").strip() and not (body.chassis_no or "").strip():
        raise HTTPException(status_code=400, detail="Thiếu Số máy/SN hoặc Số khung")
    dup = await _find_machine_by_identity(body.serial_no, body.chassis_no)
    if dup:
        raise HTTPException(status_code=400, detail="Máy đã tồn tại (Số máy/SN hoặc Số khung trùng)")

    now = _now_iso()
    code = await _next_machine_code()
    # FN-03: nếu chọn chủ sở hữu, lấy tên từ hồ sơ chủ sở hữu
    owner_name = body.owner_name
    if body.owner_id:
        owner = await db.owners.find_one({"id": body.owner_id})
        if owner:
            owner_name = owner.get("name", owner_name)
    status = body.status or "hoat_dong"
    status_source = "no_data" if status == "chua_co_du_lieu" else "manual"
    m = Machine(
        code=code, htx_id=body.htx_id,
        owner_id=body.owner_id,
        owner_name=(owner_name or htx.get("owner_name", "")),
        category_code=body.category_code,
        brand=body.brand, model=body.model, year_made=body.year_made,
        fuel=body.fuel, horsepower=body.horsepower, productivity=body.productivity,
        serial_no=(body.serial_no or "").strip(), chassis_no=(body.chassis_no or "").strip(),
        status=status, status_source=status_source, status_updated_at=now,
        condition_notes=body.condition_notes, owned_since=owned_since,
        active=True, created_at=now,
    ).model_dump()
    await db.machines.insert_one(m)
    # B5 + BR-07: ghi Lịch sử biến động loại "Thêm mới"
    await _add_machine_history(m["id"], code, "create", "manual", user["email"],
                               after=status, owned_since=owned_since)
    await db.system_logs.insert_one({
        "id": f"log-{now}", "actor_email": user["email"], "action": "CREATE_MACHINE",
        "detail": f"code={code}, htx={body.htx_id}, sn={m['serial_no']}", "ts": now,
    })
    return {k: v for k, v in m.items() if k != "_id"}


# Trường Admin được phép sửa tay
_EDITABLE_MACHINE_FIELDS = (
    "category_code", "owner_id", "owner_name", "brand", "model", "year_made",
    "fuel", "horsepower", "productivity", "serial_no", "chassis_no",
    "status", "condition_notes", "owned_since",
)


@api.patch("/machines/{machine_id}")
async def update_machine(machine_id: str, body: dict, user=Depends(require_admin)):
    cur = await db.machines.find_one({"id": machine_id}, {"_id": 0})
    if not cur:
        raise HTTPException(status_code=404, detail="Máy không tồn tại")

    changes = {}
    for k in _EDITABLE_MACHINE_FIELDS:
        if k in body:
            new_v = _coerce_date(body[k]) if k == "owned_since" else body[k]
            if new_v != cur.get(k):
                changes[k] = new_v
    if not changes:
        return {"ok": True, "changed": 0}

    # BR-01: nếu đổi định danh, kiểm tra trùng máy khác
    if "serial_no" in changes or "chassis_no" in changes:
        sn = changes.get("serial_no", cur.get("serial_no"))
        ch = changes.get("chassis_no", cur.get("chassis_no"))
        dup = await _find_machine_by_identity(sn, ch, exclude_id=machine_id)
        if dup:
            raise HTTPException(status_code=400, detail="Số máy/SN hoặc Số khung trùng máy khác")

    now = _now_iso()
    set_doc = dict(changes)
    if "status" in changes:                # BR-04: Admin sửa tay → nguồn [Nhập tay]
        set_doc["status_source"] = "no_data" if changes["status"] == "chua_co_du_lieu" else "manual"
        set_doc["status_updated_at"] = now
    await db.machines.update_one({"id": machine_id}, {"$set": set_doc})
    # BR-07: mỗi trường thay đổi ghi một dòng lịch sử (giá trị trước → sau)
    for k, v in changes.items():
        await _add_machine_history(
            machine_id, cur.get("code", ""),
            "status" if k == "status" else "update",
            "manual", user["email"], field=k, before=cur.get(k), after=v,
        )
    return {"ok": True, "changed": len(changes)}


@api.delete("/machines/{machine_id}")
async def deactivate_machine(machine_id: str, user=Depends(require_admin)):
    """BR-05: vô hiệu hóa (soft-delete) thay vì xóa cứng; giữ lịch sử."""
    cur = await db.machines.find_one({"id": machine_id}, {"_id": 0})
    if not cur:
        raise HTTPException(status_code=404, detail="Máy không tồn tại")
    if not cur.get("active", True):
        return {"ok": True, "already_inactive": True}
    now = _now_iso()
    await db.machines.update_one({"id": machine_id},
                                 {"$set": {"active": False, "deactivated_at": now}})
    # BR-08: ghi Ngày vô hiệu hóa vào lịch sử
    await _add_machine_history(machine_id, cur.get("code", ""), "deactivate",
                               "manual", user["email"], deactivated_at=now)
    return {"ok": True}


@api.post("/machines/{machine_id}/reactivate")
async def reactivate_machine(machine_id: str, user=Depends(require_admin)):
    cur = await db.machines.find_one({"id": machine_id}, {"_id": 0})
    if not cur:
        raise HTTPException(status_code=404, detail="Máy không tồn tại")
    now = _now_iso()
    await db.machines.update_one({"id": machine_id},
                                 {"$set": {"active": True, "deactivated_at": None}})
    await _add_machine_history(machine_id, cur.get("code", ""), "reactivate",
                               "manual", user["email"])
    return {"ok": True}


@api.get("/machines/{machine_id}/history")
async def machine_history(machine_id: str, user=Depends(get_current_user)):
    """B8: xem toàn bộ Lịch sử biến động của một máy (Cục & Admin đều xem được)."""
    return await db.machine_history.find(
        {"machine_id": machine_id}, {"_id": 0}
    ).sort("ts", 1).to_list(1000)


@api.get("/htx/{htx_id}/active-count")
async def htx_active_count_at(
    htx_id: str, at: Optional[str] = None, category: Optional[str] = None,
    user=Depends(get_current_user),
):
    """BR-09: số máy của HTX tại thời điểm T = đếm máy có Ngày sở hữu ≤ T và
    (chưa vô hiệu hóa hoặc Ngày vô hiệu hóa > T). Phục vụ đối chiếu biến động."""
    t = at or _now_iso()
    q = {"htx_id": htx_id, "owned_since": {"$lte": t}}
    if category and category != "ALL":
        q["category_code"] = category
    machines = await db.machines.find(q, {"_id": 0}).to_list(20000)
    in_service = [m for m in machines
                  if (m.get("deactivated_at") is None) or (m.get("deactivated_at") > t)]
    by_cat: dict = {}
    for m in in_service:
        by_cat[m["category_code"]] = by_cat.get(m["category_code"], 0) + 1
    return {"htx_id": htx_id, "at": t, "count": len(in_service), "by_category": by_cat}


# ============ OWNERS (FN-03) ============
@api.get("/owners")
async def list_owners(
    htx_id: Optional[str] = None,
    include_inactive: bool = False,
    user=Depends(get_current_user),
):
    q = {}
    if not include_inactive:
        q["active"] = {"$ne": False}
    if htx_id:
        q["htx_id"] = htx_id
    owners = await db.owners.find(q, {"_id": 0}).to_list(5000)
    # Số máy sở hữu (chỉ máy còn active)
    counts: dict = {}
    async for row in db.machines.aggregate([
        {"$match": dict(ACTIVE_FILTER)},
        {"$group": {"_id": "$owner_id", "n": {"$sum": 1}}},
    ]):
        counts[row["_id"]] = row["n"]
    for o in owners:
        o["machine_count"] = counts.get(o["id"], 0)
    return owners


@api.post("/owners")
async def create_owner(body: OwnerCreate, user=Depends(require_admin)):
    if not (body.name or "").strip():
        raise HTTPException(status_code=400, detail="Thiếu tên chủ sở hữu")
    code = await _next_owner_code()
    o = Owner(code=code, **body.model_dump()).model_dump()
    await db.owners.insert_one(o)
    return {k: v for k, v in o.items() if k != "_id"}


@api.patch("/owners/{owner_id}")
async def update_owner(owner_id: str, body: dict, user=Depends(require_admin)):
    allowed = {k: v for k, v in body.items() if k in ("name", "owner_type", "phone", "htx_id")}
    if not allowed:
        return {"ok": True}
    await db.owners.update_one({"id": owner_id}, {"$set": allowed})
    # Đồng bộ tên denormalized trên hồ sơ máy
    if "name" in allowed:
        await db.machines.update_many({"owner_id": owner_id}, {"$set": {"owner_name": allowed["name"]}})
    return {"ok": True}


@api.delete("/owners/{owner_id}")
async def deactivate_owner(owner_id: str, user=Depends(require_admin)):
    await db.owners.update_one({"id": owner_id}, {"$set": {"active": False}})
    return {"ok": True}


@api.post("/owners/{owner_id}/reactivate")
async def reactivate_owner(owner_id: str, user=Depends(require_admin)):
    await db.owners.update_one({"id": owner_id}, {"$set": {"active": True}})
    return {"ok": True}


# ============ FN-07 MAP & VISUALIZATION ============
STATUS_LABEL = {
    "hoat_dong": "Hoạt động", "bao_tri": "Bảo trì", "hong": "Hỏng",
    "chua_co_du_lieu": "Chưa có dữ liệu",
}


async def _map_pilot_categories() -> list:
    """BR-06: only categories in active, pilot-enabled stages affect the map."""
    stages = await db.stages.find({"active": {"$ne": False}}, {"_id": 0}).to_list(50)
    pilot_stages = {stage["code"] for stage in stages if stage.get("pilot_enabled", True)}
    categories = await db.machine_categories.find({"active": {"$ne": False}}, {"_id": 0}).to_list(200)
    return [category for category in categories if category.get("stage") in pilot_stages]


def _coverage_status(ratio: Optional[float], thresholds: dict) -> tuple[str, str]:
    """BR-01: classify coverage using the configured thresholds, including surplus."""
    if ratio is None:
        return "gray", "Chưa có dữ liệu"
    # Existing threshold documents may predate the surplus threshold.
    if ratio >= thresholds.get("excess_min", 1.20):
        return "blue", "Thừa"
    if ratio >= thresholds.get("sufficient_min", 0.85):
        return "green", "Đủ"
    if ratio >= thresholds.get("slight_min", 0.60):
        return "amber", "Cần chú ý"
    return "red", "Thiếu"


@api.get("/map/htx-summary")
async def map_htx_summary(
    province: Optional[str] = None,
    category: Optional[str] = None,
    status: Optional[str] = None,
    season: Optional[str] = None,
    q: Optional[str] = None,
    user=Depends(get_current_user),
):
    """Return list of HTX with computed status color for map markers."""
    htx_query = {"active": True}
    if province and province != "ALL":
        htx_query["province_code"] = province
    htx_list = await db.htx.find(htx_query, {"_id": 0}).to_list(2000)

    norms_docs = await db.productivity_norms.find({"active": {"$ne": False}}, {"_id": 0}).to_list(50)
    norms = {n["category_code"]: n["ha_per_machine_per_season"] for n in norms_docs}
    thr = await db.alert_thresholds.find_one({"key": "default"}) or {"sufficient_min": 0.85, "slight_min": 0.60, "excess_min": 1.20}
    factor = SEASON_FACTOR.get(season or "DX", 1.0)

    m_query = dict(ACTIVE_FILTER)          # BR-03: bản đồ không tính máy đã vô hiệu hóa
    if category and category != "ALL":
        m_query["category_code"] = category
    if status and status != "ALL":
        m_query["status"] = status

    all_machines = await db.machines.find(m_query, {"_id": 0}).to_list(20000)
    by_htx: dict = {}
    for m in all_machines:
        by_htx.setdefault(m["htx_id"], []).append(m)

    pilot_categories = await _map_pilot_categories()
    pilot_codes = {category["code"] for category in pilot_categories}
    category_names = {category["code"]: category["name"] for category in pilot_categories}
    result = []
    for h in htx_list:
        machines = by_htx.get(h["id"], [])
        active = [x for x in machines if x["status"] == "hoat_dong"]
        # Compute weakest ratio across categories (or filtered category)
        cats = [category] if category and category != "ALL" else list(pilot_codes)
        cats = [code for code in cats if code in pilot_codes]
        min_ratio = None
        effective_area = h.get("cultivated_area_ha", 0) * factor
        for c in cats:
            norm = norms.get(c, 0)
            if not norm or not effective_area:
                continue
            needed = effective_area / norm
            have = len([x for x in active if x["category_code"] == c])
            ratio = have / needed if needed > 0 else 1.0
            if min_ratio is None or ratio < min_ratio:
                min_ratio = ratio
        if min_ratio is None:
            color = "gray"; label = "Chưa có dữ liệu"
        elif min_ratio >= thr["sufficient_min"]:
            color = "green"; label = "Đủ"
        elif min_ratio >= thr["slight_min"]:
            color = "amber"; label = "Thiếu nhẹ"
        else:
            color = "red"; label = "Thiếu nghiêm trọng"

        # Override the legacy labels with the complete FN-07 threshold model.
        color, label = _coverage_status(min_ratio, thr)
        haystack = " ".join([
            h.get("code", ""), h.get("name", ""), h.get("owner_name", ""), h.get("commune", ""),
            *[" ".join((machine.get("code", ""), machine.get("serial_no", ""), machine.get("chassis_no", ""),
                          machine.get("owner_name", ""), category_names.get(machine.get("category_code"), ""))) for machine in machines],
        ]).lower()
        if q and q.strip().lower() not in haystack:
            continue
        result.append({
            **h,
            "machine_count": len(machines),
            "active_count": len(active),
            "coverage_ratio": round(min_ratio, 3) if min_ratio is not None else None,
            "status_color": color,
            "status_label": label,
            "pilot_category_count": len(cats),
        })
    return result


@api.get("/map/htx/{htx_id}/detail")
async def htx_detail(
    htx_id: str,
    season: Optional[str] = None,
    category: Optional[str] = None,
    status: Optional[str] = None,
    user=Depends(get_current_user),
):
    # BR-03: inactive HTX records must not be reachable from the read-only map.
    h = await db.htx.find_one({"id": htx_id, "active": True}, {"_id": 0})
    if not h:
        raise HTTPException(status_code=404, detail="HTX không tồn tại")
    machine_query = {"htx_id": htx_id, **ACTIVE_FILTER}
    if category and category != "ALL":
        machine_query["category_code"] = category
    if status and status != "ALL":
        machine_query["status"] = status
    machines = await db.machines.find(machine_query, {"_id": 0}).to_list(2000)
    cats = await _map_pilot_categories()
    norms_docs = await db.productivity_norms.find({"active": {"$ne": False}}, {"_id": 0}).to_list(50)
    norms = {n["category_code"]: n["ha_per_machine_per_season"] for n in norms_docs}

    by_cat = []
    for c in cats:
        cms = [m for m in machines if m["category_code"] == c["code"]]
        active = [m for m in cms if m["status"] == "hoat_dong"]
        norm = norms.get(c["code"], 0)
        needed = (h["cultivated_area_ha"] / norm) if norm else 0
        by_cat.append({
            "category_code": c["code"],
            "category_name": c["name"],
            "stage": c["stage"],
            "have": len(cms),
            "active": len(active),
            "needed": round(needed, 2),
            "coverage": round(len(active) / needed, 3) if needed > 0 else None,
        })
    for machine in machines:
        machine["status_label"] = STATUS_LABEL.get(machine.get("status"), "Chưa có dữ liệu")
        machine["status_source_label"] = SOURCE_LABEL.get(machine.get("status_source", "no_data"), "[Chưa có dữ liệu]")
    return {
        "htx": h,
        "machines": machines,
        "by_category": by_cat,
        "season": season,
        "area_note": "Diện tích đang hiển thị từ hồ sơ HTX; cần dữ liệu diện tích theo vụ của FN-05 để cân đối theo vụ.",
    }


# ============ DASHBOARD ============
@api.get("/dashboard/kpi")
async def dashboard_kpi(user=Depends(get_current_user)):
    total_htx = await db.htx.count_documents({"active": True})
    total_machines = await db.machines.count_documents(ACTIVE_FILTER)
    active_machines = await db.machines.count_documents({"status": "hoat_dong", **ACTIVE_FILTER})
    htx_list = await db.htx.find({"active": True}, {"_id": 0}).to_list(2000)
    total_area = sum(h.get("cultivated_area_ha", 0) for h in htx_list)

    # coverage evaluation
    norms_docs = await db.productivity_norms.find({}, {"_id": 0}).to_list(50)
    norms = {n["category_code"]: n["ha_per_machine_per_season"] for n in norms_docs}
    thr = await db.alert_thresholds.find_one({"key": "default"}) or {"sufficient_min": 0.85, "slight_min": 0.60, "excess_min": 1.20}

    machines = await db.machines.find({"status": "hoat_dong", **ACTIVE_FILTER}, {"_id": 0}).to_list(20000)
    by_htx: dict = {}
    for m in machines:
        by_htx.setdefault(m["htx_id"], []).append(m)

    sufficient = shortage_slight = shortage_severe = no_data = 0
    for h in htx_list:
        cats = [c["code"] for c in MACHINE_CATEGORIES]
        ratios = []
        for c in cats:
            norm = norms.get(c, 0)
            if not norm or not h.get("cultivated_area_ha"):
                continue
            needed = h["cultivated_area_ha"] / norm
            have = len([m for m in by_htx.get(h["id"], []) if m["category_code"] == c])
            if needed > 0:
                ratios.append(have / needed)
        if not ratios:
            no_data += 1
            continue
        mn = min(ratios)
        if mn >= thr["sufficient_min"]:
            sufficient += 1
        elif mn >= thr["slight_min"]:
            shortage_slight += 1
        else:
            shortage_severe += 1

    return {
        "total_htx": total_htx,
        "total_machines": total_machines,
        "active_machines": active_machines,
        "total_area_ha": round(total_area, 2),
        "sufficient_htx": sufficient,
        "shortage_slight_htx": shortage_slight,
        "shortage_severe_htx": shortage_severe,
        "no_data_htx": no_data,
    }


@api.get("/dashboard/charts")
async def dashboard_charts(user=Depends(get_current_user)):
    # Machines by category (chỉ máy còn hoạt động trong hệ thống)
    machines = await db.machines.find(ACTIVE_FILTER, {"_id": 0}).to_list(20000)
    cats = await db.machine_categories.find({}, {"_id": 0}).to_list(50)
    by_cat = []
    for c in cats:
        cnt = len([m for m in machines if m["category_code"] == c["code"]])
        by_cat.append({"name": c["name"], "value": cnt, "code": c["code"]})

    # Machines by province
    htx_docs = await db.htx.find({}, {"_id": 0}).to_list(2000)
    htx_prov = {h["id"]: h["province_code"] for h in htx_docs}
    provinces = await db.provinces.find({}, {"_id": 0}).to_list(50)
    by_prov = []
    for p in provinces:
        cnt = len([m for m in machines if htx_prov.get(m["htx_id"]) == p["code"]])
        area = sum(h.get("cultivated_area_ha", 0) for h in htx_docs if h["province_code"] == p["code"])
        by_prov.append({
            "province": p["name"],
            "machines": cnt,
            "area": round(area, 1),
            "density": round(cnt / area * 100, 2) if area > 0 else 0,
        })

    # HP density (HP per ha) per province
    hp_by_prov = []
    for p in provinces:
        total_hp = sum(m.get("horsepower", 0) for m in machines if htx_prov.get(m["htx_id"]) == p["code"])
        area = sum(h.get("cultivated_area_ha", 0) for h in htx_docs if h["province_code"] == p["code"])
        hp_by_prov.append({
            "province": p["name"],
            "hp_per_ha": round(total_hp / area, 3) if area > 0 else 0,
        })

    # Status distribution
    status_dist = [
        {"name": "Hoạt động", "value": len([m for m in machines if m["status"] == "hoat_dong"])},
        {"name": "Bảo trì", "value": len([m for m in machines if m["status"] == "bao_tri"])},
        {"name": "Hỏng", "value": len([m for m in machines if m["status"] == "hong"])},
    ]

    return {
        "by_category": by_cat,
        "by_province": by_prov,
        "hp_density": hp_by_prov,
        "status_distribution": status_dist,
    }


@api.get("/dashboard/priority-list")
async def priority_list(user=Depends(get_current_user)):
    """Return list of HTX with severe shortage."""
    summary = await map_htx_summary(user=user)
    critical = [h for h in summary if h["status_color"] == "red"]
    critical.sort(key=lambda x: x.get("coverage_ratio") or 1.0)
    return critical[:20]


# ============ SUPPLY-DEMAND ============
@api.get("/supply-demand")
async def supply_demand(
    province: Optional[str] = None,
    season: Optional[str] = "DX",
    user=Depends(get_current_user),
):
    factor = SEASON_FACTOR.get(season or "DX", 1.0)
    htx_q = {"active": True}
    if province and province != "ALL":
        htx_q["province_code"] = province
    htx_docs = await db.htx.find(htx_q, {"_id": 0}).to_list(2000)
    cats_docs = await db.machine_categories.find({}, {"_id": 0}).to_list(50)
    norms_docs = await db.productivity_norms.find({}, {"_id": 0}).to_list(50)
    norms = {n["category_code"]: n["ha_per_machine_per_season"] for n in norms_docs}
    thr = await db.alert_thresholds.find_one({"key": "default"}) or {"sufficient_min": 0.85, "slight_min": 0.60, "excess_min": 1.20}

    htx_ids = [h["id"] for h in htx_docs]
    all_machines = await db.machines.find({"htx_id": {"$in": htx_ids}, "status": "hoat_dong", **ACTIVE_FILTER}, {"_id": 0}).to_list(20000)

    stages_rows = []
    provinces = sorted(set(h["province_code"] for h in htx_docs))
    prov_names = {p["code"]: p["name"] for p in await db.provinces.find({}, {"_id": 0}).to_list(50)}

    for cat in cats_docs:
        norm = norms.get(cat["code"], 0)
        for pcode in provinces:
            p_htx = [h for h in htx_docs if h["province_code"] == pcode]
            area = sum(h["cultivated_area_ha"] for h in p_htx) * factor
            needed = area / norm if norm else 0
            have = len([m for m in all_machines if m["category_code"] == cat["code"] and any(h["id"] == m["htx_id"] for h in p_htx)])
            diff = have - needed
            ratio = have / needed if needed > 0 else None
            if ratio is None:
                status = "no_data"; label = "Chưa có dữ liệu"
            elif ratio >= thr.get("excess_min", 1.20):
                status = "surplus"; label = f"Thừa {int(round(diff))}"
            elif ratio >= thr["sufficient_min"]:
                status = "ok"; label = "Đủ"
            elif ratio >= thr["slight_min"]:
                status = "slight"; label = f"Thiếu {int(round(-diff))}"
            else:
                status = "severe"; label = f"Thiếu {int(round(-diff))}"
            stages_rows.append({
                "province_code": pcode,
                "province_name": prov_names.get(pcode, pcode),
                "stage": cat["stage"],
                "category_code": cat["code"],
                "category_name": cat["name"],
                "cultivated_area_ha": round(area, 1),
                "needed": round(needed, 1),
                "have": have,
                "diff": round(diff, 1),
                "coverage": round(ratio, 3) if ratio is not None else None,
                "status": status,
                "label": label,
            })
    critical = [r for r in stages_rows if r["status"] == "severe"]
    critical.sort(key=lambda x: x["coverage"] or 1.0)
    return {"rows": stages_rows, "critical": critical[:10]}


# ============ REPORTS EXPORT ============
def _make_stream(data: bytes, filename: str, content_type: str):
    bio = BytesIO(data)
    return StreamingResponse(
        bio, media_type=content_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@api.get("/reports/summary-by-region")
async def report_summary_by_region(
    province: Optional[str] = None,
    category: Optional[str] = None,
    user=Depends(get_current_user),
):
    provinces = await db.provinces.find({}, {"_id": 0}).to_list(50)
    if province and province != "ALL":
        provinces = [p for p in provinces if p["code"] == province]
    htx_docs = await db.htx.find({"active": True}, {"_id": 0}).to_list(2000)
    m_query = {}
    if category and category != "ALL":
        m_query["category_code"] = category
    machines = await db.machines.find(m_query, {"_id": 0}).to_list(20000)
    category_docs = await db.machine_categories.find({}, {"_id": 0}).to_list(500)
    category_names = {item["code"]: item["name"] for item in category_docs}
    rows = []
    for p in provinces:
        p_htx = [h for h in htx_docs if h["province_code"] == p["code"]]
        htx_ids = [h["id"] for h in p_htx]
        p_machines = [m for m in machines if m["htx_id"] in htx_ids]
        active = [m for m in p_machines if m["status"] == "hoat_dong"]
        area = sum(h["cultivated_area_ha"] for h in p_htx)
        htx_details = []
        for htx in sorted(p_htx, key=lambda item: (item.get("name", ""), item.get("code", ""))):
            htx_machines = [m for m in p_machines if m["htx_id"] == htx["id"]]
            groups = {}
            for machine in htx_machines:
                group_key = (
                    machine.get("category_code", ""),
                    machine.get("brand", ""),
                    machine.get("model", ""),
                )
                groups.setdefault(group_key, []).append(machine)
            for (category_code, brand, model), group in sorted(groups.items()):
                htx_details.append({
                    "htx_code": htx.get("code", ""),
                    "htx_name": htx.get("name", ""),
                    "htx_machine_count": len(htx_machines),
                    "category": category_names.get(category_code, category_code),
                    "brand_model": " ".join(part for part in (brand, model) if part) or "—",
                    "machine_count": len(group),
                    "active_count": sum(1 for machine in group if machine.get("status") == "hoat_dong"),
                })
        rows.append({
            "province": p["name"],
            "htx_count": len(p_htx),
            "machine_count": len(p_machines),
            "active_count": len(active),
            "area_ha": round(area, 1),
            "htx_details": htx_details,
        })
    return rows


async def _build_report(
    kind: str,
    season: Optional[str],
    province: Optional[str],
    category: Optional[str],
    user: dict,
) -> dict:
    """Shared builder that returns {title, headers, rows, filter_suffix}."""
    season_names = {s["code"]: s["name"] for s in SEASONS}
    prov_names = {p["code"]: p["name"] for p in PROVINCES}
    cat_names = {c["code"]: c["name"] for c in MACHINE_CATEGORIES}
    filter_parts = []
    if season and season != "ALL":
        filter_parts.append(f"Mùa vụ: {season_names.get(season, season)}")
    if province and province != "ALL":
        filter_parts.append(f"Tỉnh: {prov_names.get(province, province)}")
    if category and category != "ALL":
        filter_parts.append(f"Chủng loại: {cat_names.get(category, category)}")
    filter_suffix = " · ".join(filter_parts) if filter_parts else "Toàn bộ dữ liệu"

    if kind == "summary_by_region":
        rows_data = await report_summary_by_region(province=province, category=category, user=user)
        headers = [
            "Cấp dữ liệu", "Tỉnh", "Mã HTX", "Tên HTX", "Số HTX", "Tổng máy",
            "Máy hoạt động", "Diện tích (ha)", "Chủng loại máy", "Hãng / Model", "Số lượng",
        ]
        rows = []
        region_groups = []
        detail_headers = [
            "Mã HTX", "Tên HTX", "Tổng máy HTX", "Máy hoạt động",
            "Chủng loại máy", "Hãng / Model", "Số lượng",
        ]
        for region in rows_data:
            rows.append([
                "Tổng hợp tỉnh", region["province"], "", "", region["htx_count"],
                region["machine_count"], region["active_count"], region["area_ha"], "", "", "",
            ])
            detail_rows = []
            for detail in region["htx_details"]:
                detail_row = [
                    detail["htx_code"], detail["htx_name"], detail["htx_machine_count"],
                    detail["active_count"], detail["category"], detail["brand_model"], detail["machine_count"],
                ]
                detail_rows.append(detail_row)
                rows.append([
                    "Chi tiết HTX", region["province"], detail["htx_code"], detail["htx_name"], "",
                    detail["htx_machine_count"], detail["active_count"], "", detail["category"],
                    detail["brand_model"], detail["machine_count"],
                ])
            region_groups.append({
                "province": region["province"],
                "summary": (
                    f"Số HTX: {region['htx_count']} · Tổng máy: {region['machine_count']} · "
                    f"Máy hoạt động: {region['active_count']} · Diện tích: {region['area_ha']} ha"
                ),
                "rows": detail_rows,
            })
        title = f"Báo cáo Tổng hợp theo Khu vực — {filter_suffix}"
    elif kind == "supply_demand":
        sd = await supply_demand(province=province, season=season, user=user)
        rows_all = sd["rows"]
        if category and category != "ALL":
            rows_all = [r for r in rows_all if r["category_code"] == category]
        stage_names = {s["code"]: s["name"] for s in STAGES}
        headers = ["Tỉnh", "Khâu", "Chủng loại máy", "Diện tích (ha)", "Nhu cầu", "Sẵn có", "Chênh lệch", "Trạng thái"]
        rows = [[
            r["province_name"], stage_names.get(r["stage"], r["stage"]),
            r["category_name"], r["cultivated_area_ha"], r["needed"],
            r["have"], r["diff"], r["label"]
        ] for r in rows_all]
        title = f"Báo cáo Cân đối Cung – Cầu Máy móc — {filter_suffix}"
    else:  # htx_shortage
        summary = await map_htx_summary(province=province, category=category, season=season, user=user)
        shortages = [h for h in summary if h["status_color"] in ("red", "amber")]
        headers = ["Mã HTX", "Tên HTX", "Tỉnh", "Chủ sở hữu", "Diện tích", "Tổng máy", "Tỷ lệ đáp ứng", "Trạng thái"]
        rows = [[
            h["code"], h["name"], prov_names.get(h["province_code"], h["province_code"]),
            h["owner_name"], h["cultivated_area_ha"], h["machine_count"],
            f"{(h['coverage_ratio'] or 0) * 100:.1f}%", h["status_label"],
        ] for h in shortages]
        title = f"Báo cáo HTX Thừa/Thiếu Máy móc — {filter_suffix}"
    result = {"title": title, "headers": headers, "rows": rows, "filter_suffix": filter_suffix}
    if kind == "summary_by_region":
        result["region_groups"] = region_groups
        result["detail_headers"] = detail_headers
    return result


@api.get("/reports/preview")
async def report_preview(
    kind: str = Query(..., pattern="^(summary_by_region|supply_demand|htx_shortage)$"),
    season: Optional[str] = Query("DX"),
    province: Optional[str] = Query(None),
    category: Optional[str] = Query(None),
    user=Depends(get_current_user),
):
    """Return JSON preview of the report that would be exported."""
    data = await _build_report(kind, season, province, category, user)
    return {**data, "total_rows": len(data["rows"])}


@api.get("/reports/export")
async def export_report(
    kind: str = Query(..., pattern="^(summary_by_region|supply_demand|htx_shortage)$"),
    fmt: str = Query(..., pattern="^(xlsx|pdf)$"),
    season: Optional[str] = Query("DX"),
    province: Optional[str] = Query(None),
    category: Optional[str] = Query(None),
    user=Depends(get_current_user),
):
    data = await _build_report(kind, season, province, category, user)
    filename_parts = [kind]
    for x in (season, province, category):
        if x and x != "ALL":
            filename_parts.append(x)
    fname_base = "-".join(filename_parts)
    if kind == "summary_by_region":
        if fmt == "xlsx":
            blob = build_excel_by_region(data["title"], data["detail_headers"], data["region_groups"])
            return _make_stream(blob, f"{fname_base}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        blob = build_pdf_by_region(data["title"], data["detail_headers"], data["region_groups"])
        return _make_stream(blob, f"{fname_base}.pdf", "application/pdf")
    if fmt == "xlsx":
        blob = build_excel(data["title"], data["headers"], data["rows"])
        return _make_stream(blob, f"{fname_base}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        blob = build_pdf(data["title"], data["headers"], data["rows"])
        return _make_stream(blob, f"{fname_base}.pdf", "application/pdf")


# ============ ADMIN: USERS ============
@api.get("/admin/users")
async def list_users(user=Depends(require_admin)):
    docs = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(200)
    return docs


@api.post("/admin/users")
async def create_user(body: UserCreate, user=Depends(require_admin)):
    if await db.users.find_one({"email": body.email}):
        raise HTTPException(status_code=400, detail="Email đã tồn tại")
    doc = {
        "id": body.email,
        "email": body.email,
        "full_name": body.full_name,
        "role": body.role,
        "active": True,
        "must_change_password": True,
        "password_hash": hash_password(body.password),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.users.insert_one(doc)
    return {k: v for k, v in doc.items() if k not in ("_id", "password_hash")}


@api.patch("/admin/users/{email}")
async def update_user(email: str, body: dict, user=Depends(require_admin)):
    allowed = {}
    if "full_name" in body: allowed["full_name"] = body["full_name"]
    if "role" in body: allowed["role"] = body["role"]
    if "active" in body: allowed["active"] = bool(body["active"])
    if "reset_password" in body and body["reset_password"]:
        allowed["password_hash"] = hash_password(body["reset_password"])
        allowed["must_change_password"] = True
    await db.users.update_one({"email": email}, {"$set": allowed})
    return {"ok": True}


# ============ SYNC LOGS - REAL HTTP INTEGRATION ============
DEFAULT_SYNC_URL = None  # resolved at runtime to internal mock endpoint

async def _get_sync_url() -> str:
    """Return configured HTX sync URL, or fall back to internal mock endpoint."""
    doc = await db.system_settings.find_one({"key": "htx_sync_url"}, {"_id": 0})
    if doc and doc.get("value"):
        return doc["value"]
    # Default to internal mock endpoint (self-reference)
    port = os.environ.get("BACKEND_INTERNAL_PORT", "8001")
    return f"http://localhost:{port}/api/mock-htx-app/machine-updates"


@api.get("/mock-htx-app/machine-updates")
async def mock_htx_app_machine_updates():
    """Synthetic external HTX-app endpoint. Returns machine status updates.
    In production this URL is replaced by the real HTX cooperative app.
    """
    # Sample 40-80 random machines and produce status changes
    machines = await db.machines.aggregate([{"$sample": {"size": random.randint(40, 80)}}]).to_list(200)
    updates = []
    for m in machines:
        new_status = random.choices(
            ["hoat_dong", "bao_tri", "hong"],
            weights=[80, 15, 5], k=1
        )[0]
        note = ""
        if new_status == "bao_tri":
            note = random.choice(["Bảo trì định kỳ", "Thay dầu & lọc", "Kiểm tra hộp số"])
        elif new_status == "hong":
            note = random.choice(["Hỏng động cơ", "Hư hệ thống điện", "Chờ phụ tùng"])
        updates.append({
            "htx_id": m["htx_id"],
            "category_code": m["category_code"],
            "serial_no": m["serial_no"],
            "status": new_status,
            "condition_notes": note,
            "reported_at": datetime.now(timezone.utc).isoformat(),
        })
    return {"source": "HTX_APP_MOCK", "count": len(updates), "updates": updates}


@api.get("/admin/sync-logs")
async def sync_logs(user=Depends(require_admin)):
    return await db.sync_logs.find({}, {"_id": 0}).sort("started_at", -1).to_list(100)


@api.post("/admin/sync-logs/trigger")
async def trigger_sync(user=Depends(require_admin)):
    """Real HTTP integration: fetches machine status updates from configured
    HTX App endpoint and applies them to the local DB."""
    started_at = datetime.now(timezone.utc)
    started_iso = started_at.isoformat()
    sync_url = await _get_sync_url()
    updated = notfound = 0
    status = "success"
    err_msg = ""

    try:
        async with httpx.AsyncClient(timeout=15.0) as http:
            resp = await http.get(sync_url)
            resp.raise_for_status()
            payload = resp.json()
            updates = payload.get("updates", [])
            for u in updates:
                # BR-01: đối chiếu theo định danh (Số máy/SN, fallback Số khung)
                cur = await _find_machine_by_identity(
                    u.get("serial_no", ""), u.get("chassis_no", "")
                )
                if not cur or not cur.get("active", True):   # BR-11: bỏ máy đã vô hiệu hóa
                    notfound += 1
                    continue
                new_status = u.get("status", "hoat_dong")
                now_u = _now_iso()
                await db.machines.update_one({"id": cur["id"]}, {"$set": {
                    "status": new_status,
                    "condition_notes": u.get("condition_notes", ""),
                    "status_source": "app_htx",      # BR-04: nguồn [App HTX]
                    "status_updated_at": now_u,
                }})
                # BR-07: ghi lịch sử khi tình trạng thay đổi (nguồn app_htx)
                if new_status != cur.get("status"):
                    await _add_machine_history(
                        cur["id"], cur.get("code", ""), "status", "app_htx",
                        "HTX_APP", field="status",
                        before=cur.get("status"), after=new_status,
                    )
                updated += 1
            records = len(updates)
    except Exception as e:
        status = "failed"
        err_msg = str(e)[:250]
        records = 0

    finished_at = datetime.now(timezone.utc)
    latency_ms = int((finished_at - started_at).total_seconds() * 1000)
    entry = {
        "id": f"sync-{uuid.uuid4().hex[:12]}",
        "source": "HTX_APP",
        "source_url": sync_url,
        "status": status,
        "records_processed": records,
        "updated_count": updated,
        "notfound_count": notfound,
        "latency_ms": latency_ms,
        "message": err_msg or f"Cập nhật {updated} máy · Không tìm thấy {notfound}",
        "started_at": started_iso,
        "finished_at": finished_at.isoformat(),
    }
    await db.sync_logs.insert_one(dict(entry))
    await db.system_logs.insert_one({
        "id": f"log-{finished_at.timestamp()}",
        "actor_email": user["email"],
        "action": "TRIGGER_SYNC",
        "detail": f"url={sync_url}, updated={updated}, notfound={notfound}",
        "ts": finished_iso if (finished_iso := finished_at.isoformat()) else "",
    })
    return {k: v for k, v in entry.items() if k != "_id"}


@api.get("/admin/settings")
async def get_settings(user=Depends(require_admin)):
    """Fetch mutable system settings (currently only htx_sync_url)."""
    doc = await db.system_settings.find_one({"key": "htx_sync_url"}, {"_id": 0})
    default_url = await _get_sync_url()
    return {
        "htx_sync_url": (doc or {}).get("value", ""),
        "default_htx_sync_url": default_url,
    }


@api.patch("/admin/settings")
async def update_settings(body: dict, user=Depends(require_admin)):
    if "htx_sync_url" in body:
        url = (body["htx_sync_url"] or "").strip()
        await db.system_settings.update_one(
            {"key": "htx_sync_url"},
            {"$set": {"key": "htx_sync_url", "value": url}},
            upsert=True,
        )
    return {"ok": True}


# ============ MACHINES EXCEL IMPORT ============
# BR-06: cột bắt buộc (KHÔNG có cột Mã máy vì hệ thống tự sinh)
MACHINE_IMPORT_COLUMNS = [
    "htx_code", "category_code", "brand", "model", "year_made", "fuel",
    "horsepower", "productivity", "owner_name", "serial_no", "chassis_no",
    "owned_since", "status", "condition_notes",
]
_ALLOWED_MACHINE_STATUS = {"hoat_dong", "bao_tri", "hong", "chua_co_du_lieu"}


@api.get("/machines/import-template")
async def machines_import_template(user=Depends(require_admin)):
    wb = Workbook()
    ws = wb.active
    ws.title = "Machines_Template"
    ws.append(MACHINE_IMPORT_COLUMNS)
    ws.append(["CT-HTX01", "MC01", "Kubota", "L4508", 2022, "Diesel",
               90, 60, "", "SN-CAY-001", "KHUNG-CAY-001", "2022-03-15", "hoat_dong", ""])
    ws.append(["CT-HTX01", "MC04", "Yanmar", "AW82", 2021, "Diesel",
               120, 70, "", "SN-GAT-002", "KHUNG-GAT-002", "2021-06-01", "bao_tri", "Thay dầu định kỳ"])
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max(14, max_len + 2)
    bio = BytesIO(); wb.save(bio)
    return StreamingResponse(
        BytesIO(bio.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="machines-import-template.xlsx"'},
    )


@api.post("/machines/import-excel")
async def import_machines_excel(
    file: UploadFile = File(...),
    dry_run: bool = Query(False),
    user=Depends(require_admin),
):
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Chỉ chấp nhận tệp .xlsx")
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Tệp vượt quá 10MB")
    try:
        wb = load_workbook(BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Không đọc được tệp Excel: {e}")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Tệp trống hoặc chỉ có tiêu đề")
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    col_idx = {name: (header.index(name) if name in header else -1) for name in MACHINE_IMPORT_COLUMNS}
    required_cols = ("htx_code", "category_code", "owned_since")
    missing = [k for k in required_cols if col_idx.get(k, -1) == -1]
    if col_idx.get("serial_no", -1) == -1 and col_idx.get("chassis_no", -1) == -1:
        missing.append("serial_no|chassis_no")
    if missing:
        raise HTTPException(status_code=400, detail=f"Thiếu cột bắt buộc: {', '.join(missing)}")

    htx_map = {h["code"]: h for h in await db.htx.find({}, {"_id": 0}).to_list(2000)}
    cat_codes = {c["code"] for c in await db.machine_categories.find({}, {"_id": 0}).to_list(50)}
    ok_rows, error_rows = [], []
    seen_in_file: set = set()          # phát hiện trùng định danh NGAY TRONG file

    for r_i, r in enumerate(rows[1:], start=2):
        if all(c is None or str(c).strip() == "" for c in r):
            continue
        rec = {k: (r[idx] if 0 <= idx < len(r) else None) for k, idx in col_idx.items()}
        errors = []
        htx_code = str(rec.get("htx_code") or "").strip()
        cat_code = str(rec.get("category_code") or "").strip()
        serial_no = str(rec.get("serial_no") or "").strip()
        chassis_no = str(rec.get("chassis_no") or "").strip()

        if not htx_code:
            errors.append("Thiếu htx_code")
        elif htx_code not in htx_map:
            errors.append("htx_code không tồn tại")
        if not cat_code:
            errors.append("Thiếu category_code")
        elif cat_code not in cat_codes:              # BR-06: khớp danh mục chuẩn
            errors.append("category_code không khớp danh mục chuẩn")
        # BR-01: bắt buộc có Số máy/SN hoặc Số khung
        if not serial_no and not chassis_no:
            errors.append("Thiếu Số máy/SN và Số khung")
        # BR-08: Ngày HTX sở hữu bắt buộc
        owned_since = _coerce_date(rec.get("owned_since"))
        if not owned_since:
            errors.append("Thiếu owned_since (Ngày HTX sở hữu)")

        try:
            hp = float(rec.get("horsepower") or 0)
        except (TypeError, ValueError):
            errors.append("horsepower phải là số"); hp = 0
        try:
            prod = float(rec.get("productivity") or 0)
        except (TypeError, ValueError):
            errors.append("productivity phải là số"); prod = 0
        year_made = None
        if rec.get("year_made") not in (None, ""):
            try:
                year_made = int(float(rec.get("year_made")))
            except (TypeError, ValueError):
                errors.append("year_made phải là số")
        status_v = str(rec.get("status") or "hoat_dong").strip()
        if status_v not in _ALLOWED_MACHINE_STATUS:
            errors.append(f"status không hợp lệ ({', '.join(sorted(_ALLOWED_MACHINE_STATUS))})")

        # Trùng định danh ngay trong file (kể cả nhiều đợt gộp một file)
        ident = serial_no or chassis_no
        if ident and ident in seen_in_file:
            errors.append(f"Trùng định danh trong file: {ident}")
        elif ident:
            seen_in_file.add(ident)

        if errors:
            error_rows.append({"row": r_i, "serial_no": serial_no or chassis_no, "errors": errors})
        else:
            ok_rows.append({
                "row": r_i, "htx_code": htx_code, "category_code": cat_code,
                "brand": str(rec.get("brand") or "").strip(),
                "model": str(rec.get("model") or "").strip(),
                "year_made": year_made, "fuel": str(rec.get("fuel") or "").strip(),
                "horsepower": hp, "productivity": prod,
                "owner_name": str(rec.get("owner_name") or "").strip(),
                "serial_no": serial_no, "chassis_no": chassis_no,
                "owned_since": owned_since, "status": status_v,
                "condition_notes": str(rec.get("condition_notes") or "").strip(),
            })

    created = updated = deactivated = 0
    # BR-10: đối chiếu upsert theo định danh
    if not dry_run and ok_rows:
        seen_ident_by_htx: dict = {}     # htx_id -> set định danh xuất hiện trong file
        for rec in ok_rows:
            htx = htx_map[rec["htx_code"]]
            src_status = "no_data" if rec["status"] == "chua_co_du_lieu" else "manual"
            now_i = _now_iso()
            existing = await _find_machine_by_identity(rec["serial_no"], rec["chassis_no"])
            seen_ident_by_htx.setdefault(htx["id"], set()).update(
                x for x in (rec["serial_no"], rec["chassis_no"]) if x
            )
            payload = {
                "htx_id": htx["id"],
                "owner_name": rec["owner_name"] or htx["owner_name"],
                "category_code": rec["category_code"], "brand": rec["brand"],
                "model": rec["model"], "year_made": rec["year_made"], "fuel": rec["fuel"],
                "horsepower": rec["horsepower"], "productivity": rec["productivity"],
                "serial_no": rec["serial_no"], "chassis_no": rec["chassis_no"],
                "status": rec["status"], "status_source": src_status,
                "status_updated_at": now_i, "owned_since": rec["owned_since"],
                "condition_notes": rec["condition_notes"],
            }
            if existing:
                # (1) máy đã tồn tại → cập nhật nếu có thay đổi
                changed = [k for k, v in payload.items() if existing.get(k) != v]
                changed = [k for k in changed if k != "status_updated_at"]
                reactivated = not existing.get("active", True)
                set_doc = dict(payload)
                if reactivated:
                    set_doc["active"] = True
                    set_doc["deactivated_at"] = None
                await db.machines.update_one({"id": existing["id"]}, {"$set": set_doc})
                await _add_machine_history(
                    existing["id"], existing.get("code", ""), "import_update", "excel",
                    user["email"], field=",".join(changed) or "(không đổi)",
                )
                updated += 1
            else:
                # (2) máy mới → tạo bản ghi mới, sinh Mã máy
                code = await _next_machine_code()
                m = Machine(code=code, active=True, created_at=now_i, **payload).model_dump()
                await db.machines.insert_one(m)
                await _add_machine_history(
                    m["id"], code, "import_create", "excel", user["email"],
                    after=rec["status"], owned_since=rec["owned_since"],
                )
                created += 1

        # (3) máy đang có nhưng KHÔNG còn trong file của HTX đó → chuyển Inactive
        for htx_id, seen in seen_ident_by_htx.items():
            actives = await db.machines.find(
                {"htx_id": htx_id, **ACTIVE_FILTER}, {"_id": 0}
            ).to_list(20000)
            for m in actives:
                if m.get("serial_no") in seen or m.get("chassis_no") in seen:
                    continue
                now_d = _now_iso()
                await db.machines.update_one({"id": m["id"]},
                                             {"$set": {"active": False, "deactivated_at": now_d}})
                await _add_machine_history(m["id"], m.get("code", ""), "deactivate",
                                           "excel", user["email"], deactivated_at=now_d)
                deactivated += 1

        await db.system_logs.insert_one({
            "id": f"log-{_now_iso()}", "actor_email": user["email"],
            "action": "IMPORT_MACHINES_EXCEL",
            "detail": f"file={file.filename}, created={created}, updated={updated}, "
                      f"deactivated={deactivated}, errors={len(error_rows)}",
            "ts": _now_iso(),
        })

    return {
        "filename": file.filename,
        "total_rows": len(rows) - 1,
        "ok_count": len(ok_rows),
        "error_count": len(error_rows),
        "created": created,
        "updated": updated,
        "deactivated": deactivated,        # BR-10 nhánh (3)
        "dry_run": dry_run,
        "errors": error_rows[:200],
        "ok_preview": ok_rows[:20],
    }


# ============ FN-12: NHẬT KÝ HOẠT ĐỘNG (Giám sát & vận hành) ============
@api.get("/admin/system-logs")
async def system_logs(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    action: Optional[str] = None,
    actor_email: Optional[str] = None,
    user=Depends(require_admin),
):
    """B2: bộ lọc theo thời gian, loại sự kiện (action), tài khoản (actor_email)."""
    q: dict = {}
    if date_from or date_to:
        rng = {}
        if date_from: rng["$gte"] = date_from
        if date_to: rng["$lte"] = date_to
        q["ts"] = rng
    if action:
        q["action"] = action
    if actor_email:
        q["actor_email"] = actor_email
    return await db.system_logs.find(q, {"_id": 0}).sort("ts", -1).to_list(500)


@api.get("/admin/system-logs/meta")
async def system_logs_meta(user=Depends(require_admin)):
    """Danh sách loại sự kiện & tài khoản để đổ vào bộ lọc B2."""
    actions = await db.system_logs.distinct("action")
    actors = await db.system_logs.distinct("actor_email")
    return {"actions": sorted(actions), "actors": sorted(actors)}


@api.get("/admin/system-logs/stats")
async def system_logs_stats(user=Depends(require_admin)):
    """B5: thống kê hoạt động theo thời gian (ngày) và theo tài khoản."""
    logs = await db.system_logs.find({}, {"_id": 0}).to_list(5000)
    by_day: dict = {}
    by_actor: dict = {}
    for l in logs:
        day = str(l.get("ts", ""))[:10] or "unknown"
        by_day[day] = by_day.get(day, 0) + 1
        actor = l.get("actor_email", "unknown")
        by_actor[actor] = by_actor.get(actor, 0) + 1
    by_day_list = sorted(({"date": k, "count": v} for k, v in by_day.items()), key=lambda x: x["date"], reverse=True)[:30]
    by_actor_list = sorted(({"actor_email": k, "count": v} for k, v in by_actor.items()), key=lambda x: -x["count"])
    return {"by_day": by_day_list, "by_actor": by_actor_list, "total": len(logs)}


@api.get("/admin/sync-errors")
async def sync_errors(user=Depends(require_admin)):
    """B4: tab Lỗi đồng bộ — theo dõi các lần đồng bộ App HTX bị lỗi (BR-02 FN-12)."""
    return await db.sync_logs.find({"status": "failed"}, {"_id": 0}).sort("started_at", -1).to_list(200)


# ============ REGISTER ROUTER ============
app.include_router(api)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get("CORS_ORIGINS", "*").split(","),
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
async def on_startup():
    logs = await seed_all(db)
    logger.info(f"Seed done: {logs}")


@app.on_event("shutdown")
async def on_shutdown():
    client.close()
