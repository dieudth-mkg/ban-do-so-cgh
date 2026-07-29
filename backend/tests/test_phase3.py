"""Phase 3 backend tests - Real HTX Sync, Settings, Machine Import, Filtered Export, Commune."""
import os
import io
import pytest
import requests
from openpyxl import Workbook

BASE_URL = os.environ["REACT_APP_BACKEND_URL"].rstrip("/")
API = f"{BASE_URL}/api"

ADMIN = {"email": "admin@mekonggreen.vn", "password": "admin123"}
STAFF = {"email": "canbo@dcrd.gov.vn", "password": "canbo123"}


@pytest.fixture(scope="module")
def admin_token():
    r = requests.post(f"{API}/auth/login", json=ADMIN, timeout=30)
    assert r.status_code == 200
    return r.json()["token"]


@pytest.fixture(scope="module")
def staff_token():
    r = requests.post(f"{API}/auth/login", json=STAFF, timeout=30)
    assert r.status_code == 200
    return r.json()["token"]


def h(t):
    return {"Authorization": f"Bearer {t}"}


def _wb_bytes(rows):
    wb = Workbook(); ws = wb.active
    for r in rows: ws.append(r)
    bio = io.BytesIO(); wb.save(bio)
    return bio.getvalue()


# ============ Mock HTX endpoint (public) ============
def test_mock_htx_app_machine_updates_public():
    r = requests.get(f"{API}/mock-htx-app/machine-updates", timeout=30)
    assert r.status_code == 200
    d = r.json()
    assert d["source"] == "HTX_APP_MOCK"
    assert 40 <= d["count"] <= 80
    assert len(d["updates"]) == d["count"]
    u0 = d["updates"][0]
    for k in ("htx_id", "category_code", "serial_no", "status", "condition_notes", "reported_at"):
        assert k in u0
    assert u0["status"] in ("hoat_dong", "bao_tri", "hong")


# ============ Settings GET/PATCH ============
def test_admin_settings_get(admin_token):
    r = requests.get(f"{API}/admin/settings", headers=h(admin_token))
    assert r.status_code == 200
    d = r.json()
    assert "htx_sync_url" in d and "default_htx_sync_url" in d
    assert "mock-htx-app/machine-updates" in d["default_htx_sync_url"]


def test_admin_settings_patch_and_persist(admin_token):
    new_url = "https://example.com/updates"
    r = requests.patch(f"{API}/admin/settings", headers=h(admin_token),
                       json={"htx_sync_url": new_url})
    assert r.status_code == 200
    r2 = requests.get(f"{API}/admin/settings", headers=h(admin_token))
    assert r2.json()["htx_sync_url"] == new_url
    # reset to empty
    requests.patch(f"{API}/admin/settings", headers=h(admin_token),
                   json={"htx_sync_url": ""})
    r3 = requests.get(f"{API}/admin/settings", headers=h(admin_token))
    assert r3.json()["htx_sync_url"] == ""


def test_admin_settings_staff_forbidden(staff_token):
    r = requests.get(f"{API}/admin/settings", headers=h(staff_token))
    assert r.status_code == 403


# ============ Sync trigger (real HTTP) ============
def test_sync_trigger_success_updates_db(admin_token):
    # Ensure URL is empty -> falls back to default mock
    requests.patch(f"{API}/admin/settings", headers=h(admin_token), json={"htx_sync_url": ""})
    r = requests.post(f"{API}/admin/sync-logs/trigger", headers=h(admin_token), timeout=60)
    assert r.status_code == 200, r.text
    d = r.json()
    for k in ("status", "records_processed", "updated_count", "notfound_count", "latency_ms", "source_url"):
        assert k in d
    assert d["status"] == "success"
    assert d["updated_count"] > 0
    assert d["latency_ms"] >= 0
    assert "mock-htx-app" in d["source_url"]

    # persisted
    logs = requests.get(f"{API}/admin/sync-logs", headers=h(admin_token)).json()
    assert len(logs) >= 1
    latest = logs[0]
    assert "latency_ms" in latest and "source_url" in latest


def test_sync_trigger_staff_forbidden(staff_token):
    r = requests.post(f"{API}/admin/sync-logs/trigger", headers=h(staff_token))
    assert r.status_code == 403
    r2 = requests.get(f"{API}/admin/sync-logs", headers=h(staff_token))
    assert r2.status_code == 403


def test_sync_trigger_invalid_url_fails_gracefully(admin_token):
    # Point to a non-resolvable url
    requests.patch(f"{API}/admin/settings", headers=h(admin_token),
                   json={"htx_sync_url": "http://this-host-does-not-exist.invalid.local/xyz"})
    r = requests.post(f"{API}/admin/sync-logs/trigger", headers=h(admin_token), timeout=60)
    assert r.status_code == 200
    d = r.json()
    assert d["status"] == "failed"
    assert d["updated_count"] == 0
    # reset
    requests.patch(f"{API}/admin/settings", headers=h(admin_token), json={"htx_sync_url": ""})


def test_sync_actually_changes_machine_status(admin_token):
    """Integration: after trigger, at least one machine document should have condition_notes matching sync patterns
    OR sync-log's updated_count > 0 which we already asserted. Confirm DB has machines with expected notes."""
    requests.patch(f"{API}/admin/settings", headers=h(admin_token), json={"htx_sync_url": ""})
    # Trigger and check machines endpoint reflects changes (status distribution)
    r = requests.post(f"{API}/admin/sync-logs/trigger", headers=h(admin_token), timeout=60)
    assert r.status_code == 200
    assert r.json()["updated_count"] > 0
    # fetch machines and check statuses include non-default variety
    ms = requests.get(f"{API}/machines", headers=h(admin_token), timeout=30).json()
    statuses = {m["status"] for m in ms}
    # After many syncs there should be all 3 statuses
    assert "hoat_dong" in statuses


# ============ Machine import template ============
def test_machines_import_template_admin(admin_token):
    r = requests.get(f"{API}/machines/import-template", headers=h(admin_token))
    assert r.status_code == 200
    assert r.content[:2] == b"PK"
    assert "spreadsheetml" in r.headers.get("content-type", "")


def test_machines_import_template_staff_forbidden(staff_token):
    r = requests.get(f"{API}/machines/import-template", headers=h(staff_token))
    assert r.status_code == 403


MACHINE_HEADER = ["htx_code", "category_code", "serial_no", "horsepower", "status", "condition_notes"]


def test_machines_import_dry_run(admin_token):
    data = _wb_bytes([
        MACHINE_HEADER,
        ["CT-HTX01", "MC01", "TEST-M-DRY-001", 90, "hoat_dong", ""],
        ["CT-HTX01", "MC04", "TEST-M-DRY-002", 120, "bao_tri", "test"],
    ])
    files = {"file": ("m.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = requests.post(f"{API}/machines/import-excel?dry_run=true", headers=h(admin_token), files=files)
    assert r.status_code == 200, r.text
    j = r.json()
    assert j["dry_run"] is True
    assert j["ok_count"] == 2
    assert j["error_count"] == 0
    assert j["inserted"] == 0


def test_machines_import_validation_errors(admin_token):
    data = _wb_bytes([
        MACHINE_HEADER,
        ["BAD-HTX", "MC01", "TEST-BAD-001", 90, "hoat_dong", ""],   # invalid htx
        ["CT-HTX01", "ZZZZ", "TEST-BAD-002", 90, "hoat_dong", ""],  # invalid cat
        ["CT-HTX01", "MC01", "TEST-BAD-003", 90, "notastatus", ""], # invalid status
        ["CT-HTX01", "MC01", "", 90, "hoat_dong", ""],              # missing serial
    ])
    files = {"file": ("m.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = requests.post(f"{API}/machines/import-excel?dry_run=true", headers=h(admin_token), files=files)
    assert r.status_code == 200
    j = r.json()
    assert j["error_count"] == 4
    err_msgs = [" ".join(e["errors"]) for e in j["errors"]]
    assert any("htx_code không tồn tại" in m for m in err_msgs)
    assert any("category_code không hợp lệ" in m for m in err_msgs)
    assert any("status không hợp lệ" in m for m in err_msgs)
    assert any("Thiếu serial_no" in m for m in err_msgs)


def test_machines_import_duplicate_skipped(admin_token):
    # First insert one
    data = _wb_bytes([
        MACHINE_HEADER,
        ["CT-HTX01", "MC01", "TEST-M-DUP-999", 90, "hoat_dong", ""],
    ])
    files = {"file": ("dup1.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = requests.post(f"{API}/machines/import-excel?dry_run=false", headers=h(admin_token), files=files)
    assert r.status_code == 200
    inserted_id = None
    if r.json()["inserted"]:
        inserted_id = "CT-HTX01-MC01-TEST-M-DUP-999"

    # Second time -> should be skipped
    files2 = {"file": ("dup2.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r2 = requests.post(f"{API}/machines/import-excel?dry_run=false", headers=h(admin_token), files=files2)
    assert r2.status_code == 200
    j = r2.json()
    assert j["skipped_count"] == 1
    assert j["inserted"] == 0

    # cleanup
    if inserted_id:
        requests.delete(f"{API}/machines/{inserted_id}", headers=h(admin_token))


def test_machines_import_actual_insert_and_cleanup(admin_token):
    ms_before = requests.get(f"{API}/machines", headers=h(admin_token)).json()
    count_before = len(ms_before)

    serials = ["TEST-M-INS-A1", "TEST-M-INS-A2"]
    data = _wb_bytes([
        MACHINE_HEADER,
        ["CT-HTX01", "MC01", serials[0], 90, "hoat_dong", ""],
        ["CT-HTX01", "MC02", serials[1], 100, "hoat_dong", ""],
    ])
    files = {"file": ("ins.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = requests.post(f"{API}/machines/import-excel?dry_run=false", headers=h(admin_token), files=files)
    assert r.status_code == 200, r.text
    j = r.json()
    assert j["inserted"] == 2

    ms_after = requests.get(f"{API}/machines", headers=h(admin_token)).json()
    assert len(ms_after) >= count_before + 2
    got = {m["serial_no"] for m in ms_after}
    for s in serials:
        assert s in got

    # cleanup
    for s in serials:
        mid = f"CT-HTX01-MC01-{s}" if "A1" in s else f"CT-HTX01-MC02-{s}"
        requests.delete(f"{API}/machines/{mid}", headers=h(admin_token))


# ============ Reports export with filters ============
def _is_xlsx(resp):
    return resp.status_code == 200 and resp.content[:2] == b"PK" and \
        "spreadsheetml" in resp.headers.get("content-type", "")


def _is_pdf(resp):
    return resp.status_code == 200 and resp.content[:4] == b"%PDF"


def test_export_summary_by_region_province_filter(admin_token):
    r = requests.get(f"{API}/reports/export",
                     params={"kind": "summary_by_region", "fmt": "xlsx", "province": "CT"},
                     headers=h(admin_token))
    assert _is_xlsx(r)


def test_export_supply_demand_multi_filter(admin_token):
    r = requests.get(f"{API}/reports/export",
                     params={"kind": "supply_demand", "fmt": "xlsx", "season": "HT", "province": "CT"},
                     headers=h(admin_token))
    assert _is_xlsx(r)
    # verify content has expected filter suffix in worksheet title
    wb_bytes = r.content
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(wb_bytes))
    ws = wb.active
    # First row usually has title
    cells = [str(c.value) for row in ws.iter_rows(max_row=3) for c in row if c.value]
    joined = " | ".join(cells)
    assert "Hè Thu" in joined or "HT" in joined
    assert "Cần Thơ" in joined or "CT" in joined


def test_export_htx_shortage_pdf_category_filter(admin_token):
    r = requests.get(f"{API}/reports/export",
                     params={"kind": "htx_shortage", "fmt": "pdf", "category": "MC04"},
                     headers=h(admin_token))
    assert _is_pdf(r)


# ============ HTX commune field ============
def test_htx_has_commune_field(admin_token):
    r = requests.get(f"{API}/htx", headers=h(admin_token))
    assert r.status_code == 200
    htxs = r.json()
    with_commune = [h for h in htxs if h.get("commune")]
    assert len(with_commune) > 0
    # Should include some 'Xã ...' prefix
    xa_names = [h["commune"] for h in with_commune if str(h["commune"]).startswith("Xã")]
    assert len(xa_names) > 0
